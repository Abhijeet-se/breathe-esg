"""
Base Parser
===========

Abstract base class for all data source parsers.

Design Decisions:
- Parsers are stateless - they receive a file path and return parsed data
- Each parser defines its own expected headers and row parsing logic
- The base class handles CSV/Excel detection, file reading, and error collection
- Errors are returned alongside data (not raised) so partial results are preserved
- This enables batch-level reporting: "85 of 100 rows parsed successfully"
"""

import csv
import io
import logging
from abc import ABC, abstractmethod
from typing import List, Dict, Tuple, Any, Optional

logger = logging.getLogger(__name__)


class ParseError:
    """
    Structured error from parsing a single row.

    Attributes:
        row_number: 1-indexed row from the original file
        field: Which field caused the error (None for row-level errors)
        message: Human-readable error description
        original_data: The raw row data that caused the error
    """
    def __init__(self, row_number: int, field: Optional[str], message: str,
                 original_data: Optional[Dict] = None):
        self.row_number = row_number
        self.field = field
        self.message = message
        self.original_data = original_data or {}

    def to_dict(self) -> Dict:
        return {
            'row_number': self.row_number,
            'field': self.field,
            'message': self.message,
        }


class BaseParser(ABC):
    """
    Abstract base class for file parsers.

    Subclasses must implement:
    - parse_row(row_data, row_number) -> parsed dict or raises ValueError
    - get_expected_headers() -> list of expected column names
    - get_source_type() -> string identifier for this source type

    Usage:
        parser = SAPParser()
        results, errors = parser.parse_file(file_path_or_file_object)
    """

    @abstractmethod
    def parse_row(self, row_data: Dict[str, str], row_number: int) -> Dict[str, Any]:
        """
        Parse a single row of data into a normalized dictionary.

        Args:
            row_data: Dictionary of column_name -> value from the file
            row_number: 1-indexed row number for error reporting

        Returns:
            Dictionary with parsed and typed values

        Raises:
            ValueError: If the row cannot be parsed (with descriptive message)
        """
        pass

    @abstractmethod
    def get_expected_headers(self) -> List[str]:
        """Return list of expected column headers (in English/normalized form)."""
        pass

    @abstractmethod
    def get_source_type(self) -> str:
        """Return the source_type identifier (e.g., 'sap_fuel')."""
        pass

    def get_header_mappings(self) -> Dict[str, str]:
        """
        Return a mapping of alternative header names to standard names.
        Override in subclasses to support non-English headers.

        Returns:
            Dict mapping source_header -> standard_header
        """
        return {}

    def normalize_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Map actual file headers to expected standard headers.

        Tries exact match first, then case-insensitive match, then
        falls back to the header_mappings for alternative names (e.g., German).

        Returns:
            Dict mapping original_header -> standard_header
        """
        mappings = self.get_header_mappings()
        expected = {h.lower(): h for h in self.get_expected_headers()}
        result = {}

        for header in headers:
            header_stripped = header.strip()
            header_lower = header_stripped.lower()

            # Try exact match with expected headers
            if header_lower in expected:
                result[header_stripped] = expected[header_lower]
            # Try alternative name mappings (e.g., German -> English)
            elif header_stripped in mappings:
                result[header_stripped] = mappings[header_stripped]
            elif header_lower in {k.lower(): v for k, v in mappings.items()}:
                # Case-insensitive mapping lookup
                for k, v in mappings.items():
                    if k.lower() == header_lower:
                        result[header_stripped] = v
                        break
            else:
                # Unknown header - keep as-is (will be in original_data but not parsed)
                result[header_stripped] = header_stripped

        return result

    def parse_file(self, file_obj, file_name: str = '') -> Tuple[List[Dict], List[ParseError]]:
        """
        Parse an uploaded file (CSV or Excel) and return results + errors.

        Detects file type from extension, reads all rows, maps headers,
        and calls parse_row() for each data row.

        Args:
            file_obj: A file-like object (Django's UploadedFile or similar)
            file_name: Original filename (used for format detection)

        Returns:
            Tuple of (parsed_results, parse_errors)
            - parsed_results: List of dicts with parsed data + row_number + original_data
            - parse_errors: List of ParseError objects for rows that failed
        """
        results = []
        errors = []

        # Detect file format from extension
        name_lower = file_name.lower()
        if name_lower.endswith(('.xlsx', '.xls')):
            rows = self._read_excel(file_obj)
        elif name_lower.endswith('.csv'):
            rows = self._read_csv(file_obj)
        else:
            # Default to CSV - most common format
            logger.warning(f"Unknown file extension for '{file_name}', trying CSV")
            rows = self._read_csv(file_obj)

        if not rows:
            errors.append(ParseError(0, None, "File is empty or has no data rows"))
            return results, errors

        # First row is used to detect headers (already handled by readers)
        for row_number, row_data in enumerate(rows, start=1):
            try:
                parsed = self.parse_row(row_data, row_number)
                parsed['_row_number'] = row_number
                parsed['_original_data'] = dict(row_data)  # Preserve original
                results.append(parsed)
            except (ValueError, KeyError, TypeError) as e:
                errors.append(ParseError(
                    row_number=row_number,
                    field=None,
                    message=str(e),
                    original_data=dict(row_data),
                ))

        logger.info(
            f"Parsed {file_name}: {len(results)} successful, {len(errors)} errors"
        )
        return results, errors

    def _read_csv(self, file_obj) -> List[Dict]:
        """
        Read a CSV file and return list of dicts (one per data row).

        Handles encoding detection and header mapping.
        """
        # Reset file position if possible
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)

        # Read file content - handle both binary and text modes
        if hasattr(file_obj, 'read'):
            content = file_obj.read()
            if isinstance(content, bytes):
                # Try UTF-8 first, fall back to latin-1 (common for German SAP exports)
                try:
                    content = content.decode('utf-8')
                except UnicodeDecodeError:
                    content = content.decode('latin-1')
        else:
            content = str(file_obj)

        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            return []

        # Map headers using our normalize_headers logic
        header_map = self.normalize_headers(list(reader.fieldnames))

        rows = []
        for row in reader:
            # Remap column names to standard names
            mapped_row = {}
            for original_key, value in row.items():
                if original_key and original_key.strip() in header_map:
                    standard_key = header_map[original_key.strip()]
                    mapped_row[standard_key] = value.strip() if value else ''
                elif original_key:
                    mapped_row[original_key.strip()] = value.strip() if value else ''
            rows.append(mapped_row)

        return rows

    def _read_excel(self, file_obj) -> List[Dict]:
        """
        Read an Excel file (.xlsx) using openpyxl and return list of dicts.

        Only reads the first worksheet. Skips completely empty rows.
        """
        from openpyxl import load_workbook

        # Reset file position
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)

        wb = load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows()

        # First row = headers
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        raw_headers = [str(cell.value).strip() if cell.value else '' for cell in header_row]
        header_map = self.normalize_headers(raw_headers)

        rows = []
        for row in rows_iter:
            values = [cell.value for cell in row]
            # Skip completely empty rows
            if all(v is None or str(v).strip() == '' for v in values):
                continue

            mapped_row = {}
            for header, value in zip(raw_headers, values):
                if header and header in header_map:
                    standard_key = header_map[header]
                    mapped_row[standard_key] = str(value).strip() if value is not None else ''
                elif header:
                    mapped_row[header] = str(value).strip() if value is not None else ''
            rows.append(mapped_row)

        wb.close()
        return rows
